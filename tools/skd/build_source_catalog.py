from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DRIVE_ID_RE = re.compile(r"(?:/file/d/|[?&]id=)([A-Za-z0-9_-]+)")


def extract_drive_file_id(url: str | None) -> str:
    if not url:
        return ""
    match = DRIVE_ID_RE.search(url)
    return match.group(1) if match else ""


def cell_value(cell: Any) -> str:
    value = cell.value
    return "" if value is None else str(value).strip()


def build_catalog(xlsx_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' tidak ditemukan. Sheet tersedia: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    rows: list[dict[str, str]] = []
    section = "pusat"

    for row_number in range(2, sheet.max_row + 1):
        entity_name = cell_value(sheet.cell(row_number, 1))
        entity_type = cell_value(sheet.cell(row_number, 2))
        pdf_name = cell_value(sheet.cell(row_number, 3))

        if not entity_name and not entity_type and not pdf_name:
            continue

        if entity_name.upper() == "PEMDA" and pdf_name.upper() == "HASIL SKD":
            section = "pemda"
            continue

        if entity_name == "Kementerian/Lembaga/Koordinator" and pdf_name == "HASIL SKD":
            continue

        link = sheet.cell(row_number, 3).hyperlink
        drive_url = link.target if link else ""
        drive_file_id = extract_drive_file_id(drive_url)

        warnings: list[str] = []
        if not entity_name:
            warnings.append("missing_entity")
        if not pdf_name:
            warnings.append("missing_pdf_name")
        if not drive_url:
            warnings.append("missing_drive_link")
        if pdf_name and "2021" in pdf_name and "2024" not in pdf_name:
            warnings.append("filename_year_maybe_not_2024")

        rows.append(
            {
                "sheet_row": str(row_number),
                "section": section,
                "entity_name": entity_name,
                "entity_type": entity_type,
                "pdf_name": pdf_name,
                "drive_file_id": drive_file_id,
                "drive_url": drive_url,
                "tahun": "2024",
                "warnings": ";".join(warnings),
            }
        )

    duplicate_keys = Counter(row["drive_file_id"] or row["pdf_name"].lower() for row in rows)
    for row in rows:
        duplicate_key = row["drive_file_id"] or row["pdf_name"].lower()
        row["duplicate_count"] = str(duplicate_keys[duplicate_key])
        if duplicate_keys[duplicate_key] > 1:
            row["warnings"] = ";".join(filter(None, [row["warnings"], "duplicate_source"]))

    return rows


def write_outputs(rows: list[dict[str, str]], output_csv: Path, output_json: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "sheet_row",
        "section",
        "entity_name",
        "entity_type",
        "pdf_name",
        "drive_file_id",
        "drive_url",
        "tahun",
        "duplicate_count",
        "warnings",
    ]

    with output_csv.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    summary = {
        "total_sources": len(rows),
        "with_drive_link": sum(1 for row in rows if row["drive_url"]),
        "without_drive_link": sum(1 for row in rows if not row["drive_url"]),
        "duplicate_sources": sum(1 for row in rows if int(row["duplicate_count"]) > 1),
        "sections": dict(Counter(row["section"] for row in rows)),
        "rows": rows,
    }

    output_json.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build local SKD PDF source catalog from the Riset 2026 XLSX export.")
    parser.add_argument("--xlsx", default="data/raw/riset-2026.xlsx", help="Path to exported XLSX.")
    parser.add_argument("--sheet", default="HASIL SKD 2024", help="Sheet tab name.")
    parser.add_argument("--csv", default="data/staging/pdf_sources_catalog.csv", help="Output CSV path.")
    parser.add_argument("--json", default="data/staging/pdf_sources_catalog.json", help="Output JSON path.")
    args = parser.parse_args()

    rows = build_catalog(Path(args.xlsx), args.sheet)
    write_outputs(rows, Path(args.csv), Path(args.json))
    print(f"Built {len(rows)} sources -> {args.csv}")


if __name__ == "__main__":
    main()
